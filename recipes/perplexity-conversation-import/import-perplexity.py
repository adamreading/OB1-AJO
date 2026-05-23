#!/usr/bin/env python3
"""
Open Brain — Perplexity Export Importer

Extracts conversations and memories from a Perplexity data export (.xlsx),
filters deleted/already-imported items, summarizes conversations into
1-3 distilled thoughts via LLM, and loads everything into your Open Brain.

Memory entries are ingested directly (already summarized by Perplexity).
JSON profile rows (MEMORY_KEY empty, MEMORY_VALUE is a JSON object) are
flattened into per-section thoughts.

Usage:
    python import-perplexity.py path/to/export.xlsx [options]

Ingestion modes:
    Default:              Supabase direct insert (requires SUPABASE_URL,
                          SUPABASE_SERVICE_ROLE_KEY, OPENROUTER_API_KEY)

Options:
    --xlsx PATH           Path to Perplexity .xlsx export (required)
    --dry-run             Parse, filter, summarize, but don't ingest
    --after YYYY-MM-DD    Only conversations created after this date
    --before YYYY-MM-DD   Only conversations created before this date
    --limit N             Max items per type to process
    --type TYPE           What to import: conversations, memory, or both (default: both)
    --model MODEL         LLM backend: openrouter (default) or ollama
    --ollama-model NAME   Ollama model name (default: qwen3)
    --verbose             Show full content during processing
    --report FILE         Write a markdown report of everything imported

Environment variables:
    SUPABASE_URL               Supabase project URL
    SUPABASE_SERVICE_ROLE_KEY  Supabase service role key
    OPENROUTER_API_KEY         OpenRouter API key (summarization + embeddings)
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

# ─── Configuration ───────────────────────────────────────────────────────────

SYNC_LOG_PATH = Path("perplexity-sync-log.json")

OPENROUTER_BASE = "https://openrouter.ai/api/v1"
OLLAMA_BASE = "http://localhost:11434"

# Supabase: reads the "Secret Key" from Settings → API (starts with sb_secret_)
# The env var name uses the legacy convention for cross-recipe consistency.
SUPABASE_URL = os.environ.get("SUPABASE_URL", "") or os.environ.get("OPEN_BRAIN_URL", "")
# AJO-local: this fork stores the service-role secret as SUPABASE_KEY.
# Fall back to the upstream name and the upstream Open Brain naming
# (OPEN_BRAIN_SERVICE_KEY) too so the recipe works against any flavour.
SUPABASE_SERVICE_ROLE_KEY = (
    os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    or os.environ.get("SUPABASE_KEY", "")
    or os.environ.get("OPEN_BRAIN_SERVICE_KEY", "")
)
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", "")

# ─── Heuristic prefilter ────────────────────────────────────────────────────
# AJO-local: drop obvious trivia BEFORE we spend OpenRouter tokens on a
# summarisation that's likely to come back {"thoughts": []} anyway.
# Conservative — when in doubt the conversation goes through to the LLM
# judge, which is itself selective. Tuned for Perplexity-shaped Q&A.

JUNK_TITLE_PATTERNS = [
    r"^how many\s+",                  # "how many r's in strawberry"
    r"^what (time|date|day)\s",       # "what time is it in tokyo"
    r"^when (does|did|is)\s",         # "when does X open"
    r"^how do you spell\s",
    r"^spell\s\w+\s*$",
    r"^convert\s+[\d\.]+\s",          # "convert 32f to c"
    r"^\d+\s*(°|degrees)",            # "32 degrees fahrenheit"
    r"^\d+\s*[\+\-\*/]\s*\d+",        # "5 + 3"
    r"^(define|definition of)\s",     # "define cromulent"
    r"^translate\s",
    r"^pronounce\s",
    r"^what does (.{1,30})\s+mean\s*\??$",  # "what does X mean?" short ones
    r"^image\.(jpg|png|jpeg|webp)\s*$",  # bare image-upload prompts
]
_JUNK_TITLE_RE = [re.compile(p, re.IGNORECASE) for p in JUNK_TITLE_PATTERNS]

# Minimum answer length to bother summarising. Real Perplexity answers
# with sources and structure tend to be >400 chars; trivia is often
# <200. 400 is a defensible midpoint that catches most one-liners
# without dropping shorter-but-substantive answers.
MIN_ANSWER_CHARS = 400


def looks_like_junk(conv):
    """Return (skip: bool, reason: str). Conservative — only drop on clear
    trivia signals. Anything ambiguous goes through to the LLM judge."""
    answer = (conv.get("answer_text") or "").strip()
    title = (conv.get("title") or "").strip()

    if not answer:
        return True, "empty answer"
    if len(answer) < MIN_ANSWER_CHARS:
        return True, f"answer too short ({len(answer)} chars < {MIN_ANSWER_CHARS})"
    if not title:
        # No title + an answer that survived the length check is unusual
        # but not necessarily junk; let the LLM decide.
        return False, ""
    for rx in _JUNK_TITLE_RE:
        if rx.search(title):
            return True, f"trivia title pattern: {rx.pattern}"
    return False, ""


SUMMARIZATION_PROMPT = """\
You are distilling a Perplexity Q&A exchange into standalone thoughts for a \
personal knowledge base. Your job is to be HIGHLY SELECTIVE — only extract \
knowledge that would be valuable to retrieve months or years from now.

You will receive a search query and Perplexity's answer.

CAPTURE these (1-3 thoughts max):
- Decisions made and the reasoning behind them
- People, places, or topics explored with lasting relevance
- Lessons learned, preferences discovered, or useful frameworks
- Research findings worth remembering
- Context about the user's interests, projects, or goals

SKIP these entirely (return empty):
- Simple factual lookups (restaurant hours, definitions, recipes)
- One-off trivia with no lasting value
- Generic how-to with no personal context

Each thought must be:
- A clear, standalone statement (makes sense without the Q&A)
- Written in first person
- Anchored with names, dates, or context when available
- 1-3 sentences

Return JSON: {"thoughts": ["thought1", "thought2"]}
If nothing worth capturing, return {"thoughts": []}
Err on the side of returning empty — less is more."""

# JSON profile flattening config
PROFILE_SECTIONS = {
    "demographics": "Demographics",
    "interests": "Interests",
    "work_and_education": "Work and Education",
    "lifestyle": "Lifestyle",
    "technology": "Technology",
    "knowledge": "Knowledge and Expertise",
    "personal_traits": "Personal Traits",
}

# ─── Sync Log ────────────────────────────────────────────────────────────────


def load_sync_log():
    """Load sync log from disk. Returns dict with ingested_ids and last_sync."""
    try:
        with open(SYNC_LOG_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"ingested_ids": {}, "last_sync": ""}


def save_sync_log(log):
    """Save sync log to disk."""
    with open(SYNC_LOG_PATH, "w") as f:
        json.dump(log, f, indent=2)


def make_dedupe_key(*parts):
    """Generate a short SHA256 hash from string parts for deduplication."""
    raw = "|".join(str(p) for p in parts)
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


# ─── HTTP Helpers ────────────────────────────────────────────────────────────

try:
    import requests
except ImportError:
    print("Missing dependency: requests")
    print("Install with: pip install requests")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Missing dependency: openpyxl")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def http_post_with_retry(url, headers, body, retries=2):
    """POST with exponential backoff retry on transient failures."""
    for attempt in range(retries + 1):
        try:
            resp = requests.post(url, headers=headers, json=body, timeout=30)
            if resp.status_code >= 500 and attempt < retries:
                time.sleep(1 * (attempt + 1))
                continue
            return resp
        except requests.RequestException:
            if attempt < retries:
                time.sleep(1 * (attempt + 1))
                continue
            raise
    return None  # unreachable


# ─── XLSX Parsing ────────────────────────────────────────────────────────────


def _cell_value(cell):
    """Extract a clean string value from a cell, treating None as empty."""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


TIMESTAMP_FORMATS = (
    "%Y-%m-%d %H:%M:%S.%f",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
)


def parse_timestamp_iso(raw):
    """Parse a Perplexity timestamp string into ISO 8601 (UTC).

    Returns ISO string or None if parsing fails.
    Also handles datetime objects from openpyxl (data_only mode).
    """
    if not raw:
        return None

    # openpyxl may return datetime objects directly
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%dT%H:%M:%S+00:00")

    raw = str(raw).strip()
    if not raw:
        return None

    for fmt in TIMESTAMP_FORMATS:
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.strftime("%Y-%m-%dT%H:%M:%S+00:00")
        except ValueError:
            continue

    return None


def extract_conversations(path):
    """Extract conversation rows from any supported export format.

    Returns list of dicts with keys: uuid, created, updated, title, answer_text.

    Supported inputs (auto-detected by file extension):
      .xlsx  — workbook with a "Conversations" sheet (canonical format)
      .csv   — UUID,CREATED,UPDATED,TITLE,OUTPUT_STR
      .json  — { "conversations": [ { context_uuid, context_title,
                                       created_at, updated_at,
                                       entries: [{query, answer, ...}] } ] }
               Multi-turn entries get concatenated into a single answer_text
               for the summariser (lossier than emitting one thought per
               turn, but matches the recipe's "1-3 thoughts per conv"
               output shape, and the LLM still sees the full conversation).
    """
    ext = str(path).lower().rsplit(".", 1)[-1] if "." in str(path) else ""
    if ext == "json":
        return _extract_conversations_json(path)
    if ext == "csv":
        return _extract_conversations_csv(path)
    return _extract_conversations_xlsx(path)


def _extract_conversations_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Conversations" not in wb.sheetnames:
        print("Warning: No 'Conversations' sheet found in export.")
        wb.close()
        return []

    ws = wb["Conversations"]
    rows = list(ws.iter_rows())
    wb.close()

    if len(rows) < 2:
        return []

    header = [_cell_value(c) for c in rows[0]]
    col_idx = {name: i for i, name in enumerate(header)}

    conversations = []
    for row in rows[1:]:
        values = [_cell_value(c) for c in row]

        uuid = values[col_idx.get("UUID", -1)] if "UUID" in col_idx else ""
        created = values[col_idx.get("CREATED", -1)] if "CREATED" in col_idx else ""
        updated = values[col_idx.get("UPDATED", -1)] if "UPDATED" in col_idx else ""
        title = values[col_idx.get("TITLE", -1)] if "TITLE" in col_idx else ""
        output_str = (
            values[col_idx.get("OUTPUT_STR", -1)] if "OUTPUT_STR" in col_idx else ""
        )

        answer_text = _parse_output_str(output_str)

        if not uuid and not title:
            continue

        conversations.append(
            {
                "uuid": uuid,
                "created": created,
                "updated": updated,
                "title": title,
                "answer_text": answer_text,
            }
        )

    return conversations


def _extract_conversations_csv(csv_path):
    """CSV variant of the Conversations sheet (same headers)."""
    import csv as _csv

    conversations = []
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        for row in _csv.DictReader(f):
            uuid = (row.get("UUID") or "").strip()
            title = (row.get("TITLE") or "").strip()
            if not uuid and not title:
                continue
            conversations.append(
                {
                    "uuid": uuid,
                    "created": (row.get("CREATED") or "").strip(),
                    "updated": (row.get("UPDATED") or "").strip(),
                    "title": title,
                    "answer_text": _parse_output_str(row.get("OUTPUT_STR") or ""),
                }
            )
    return conversations


def _extract_conversations_json(json_path):
    """JSON variant — Perplexity's newer export format with multi-turn entries.

    Each conversation has an entries[] array of (query, answer) turns. We
    flatten them into a single answer_text using Q1/A1/Q2/A2/... so the
    summariser sees the full thread and can synthesise a multi-turn-aware
    thought set.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    convs = data.get("conversations", []) if isinstance(data, dict) else []
    out = []
    for c in convs:
        uuid = (c.get("context_uuid") or "").strip()
        title = (c.get("context_title") or "").strip()
        if not uuid and not title:
            continue
        entries = c.get("entries") or []
        # Flatten multi-turn entries. For single-turn we just emit the
        # bare answer (matches the xlsx/csv format) so the summariser
        # prompt isn't bloated with Q1/A1 framing for trivial cases.
        if len(entries) <= 1:
            answer_text = (entries[0].get("answer") if entries else "") or ""
        else:
            parts = []
            for i, e in enumerate(entries, start=1):
                q = (e.get("query") or "").strip()
                a = (e.get("answer") or "").strip()
                if q:
                    parts.append(f"Q{i}: {q}")
                if a:
                    parts.append(f"A{i}: {a}")
            answer_text = "\n\n".join(parts)
        out.append(
            {
                "uuid": uuid,
                "created": c.get("created_at") or "",
                "updated": c.get("updated_at") or "",
                "title": title,
                "answer_text": answer_text.strip(),
                # AJO-local: keep the Perplexity Space id + mode so the
                # ingest stage can resolve a friendly source_type tag
                # (perplexity-scl, perplexity-personal, etc.) and tuck
                # the original mode into metadata for later filtering.
                "space_uuid": c.get("collection_uuid"),
                "mode": c.get("mode"),
            }
        )
    return out


def load_space_map(path):
    """Read the optional Perplexity-spaces JSON map.

    Format produced by the AJO template-generator:
      { "<uuid>": { "name": "scl", "classification": "work",
                    "action": "import"|"drop", ... } }

    Returns {uuid: {"name": <slug>, "classification": <work|personal|"">,
                    "action": <import|drop>}} for every mapped UUID. Spaces
    with action="drop" cause every conversation in that space to be
    skipped at the prefilter stage. Spaces without a classification leave
    that column NULL so the worker classifies on entity extraction.
    """
    if not path or not os.path.isfile(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}
    out = {}
    for uuid, info in (raw or {}).items():
        if not isinstance(info, dict):
            continue
        name = (info.get("name") or "").strip()
        classification = (info.get("classification") or "").strip().lower()
        if classification not in ("work", "personal", ""):
            classification = ""
        action = (info.get("action") or "import").strip().lower()
        if action not in ("import", "drop"):
            action = "import"
        slug = ""
        if name:
            slug = re.sub(r"[^a-z0-9-]+", "-", name.lower()).strip("-")
        # Keep entries even when name/classification are blank, so the
        # action='drop' marker is honoured for spaces with empty names.
        out[uuid] = {"name": slug, "classification": classification, "action": action}
    return out


def _parse_output_str(output_str):
    """Extract answer text from Perplexity's OUTPUT_STR JSON blob."""
    if not output_str:
        return ""

    try:
        data = json.loads(output_str)
    except (json.JSONDecodeError, TypeError):
        return output_str

    if isinstance(data, dict):
        answer = data.get("answer", "")
        if isinstance(answer, str):
            return answer.strip()
        return json.dumps(answer) if answer else ""

    return output_str


def extract_memory_rows(path):
    """Extract memory rows from either a CSV file or the 'Memory' sheet of an XLSX."""
    ext = str(path).lower().rsplit(".", 1)[-1] if "." in str(path) else ""
    if ext == "csv":
        return _extract_memory_csv(path)
    return _extract_memory_xlsx(path)


def _normalize_memory_row(mem):
    """Shared normalisation for the IS_* boolean columns."""
    for bool_col in ("IS_DELETED", "IS_FORGOTTEN", "IS_INVISIBLE"):
        val = (mem.get(bool_col) or "").strip().lower() if isinstance(mem.get(bool_col), str) else mem.get(bool_col)
        if isinstance(val, str):
            mem[bool_col] = val in ("true", "1", "yes")
        else:
            mem[bool_col] = bool(val)
    return mem


def _extract_memory_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Memory" not in wb.sheetnames:
        print("Warning: No 'Memory' sheet found in export.")
        wb.close()
        return []

    ws = wb["Memory"]
    rows = list(ws.iter_rows())
    wb.close()

    if len(rows) < 2:
        return []

    header = [_cell_value(c) for c in rows[0]]
    col_idx = {name: i for i, name in enumerate(header)}

    memories = []
    for row in rows[1:]:
        values = [_cell_value(c) for c in row]

        mem = {}
        for col_name, idx in col_idx.items():
            mem[col_name] = values[idx] if idx < len(values) else ""

        _normalize_memory_row(mem)

        if not mem.get("MEMORY_KEY") and not mem.get("MEMORY_VALUE"):
            continue

        memories.append(mem)

    return memories


def _extract_memory_csv(csv_path):
    """CSV variant of the Memory sheet."""
    import csv as _csv

    memories = []
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        for row in _csv.DictReader(f):
            mem = {k: (v if v is not None else "") for k, v in row.items()}
            _normalize_memory_row(mem)
            if not mem.get("MEMORY_KEY") and not mem.get("MEMORY_VALUE"):
                continue
            memories.append(mem)
    return memories


# ─── JSON Profile Handling ──────────────────────────────────────────────────


def is_json_profile_row(row):
    """Detect if a memory row is a JSON profile (no MEMORY_KEY, MEMORY_VALUE is JSON)."""
    memory_key = row.get("MEMORY_KEY", "").strip()
    memory_value = row.get("MEMORY_VALUE", "").strip()

    if memory_key:
        return False

    if not memory_value:
        return False

    if not memory_value.startswith("{"):
        return False

    try:
        json.loads(memory_value)
        return True
    except (json.JSONDecodeError, TypeError):
        return False


def flatten_json_section(key, value):
    """Flatten a JSON section into a natural-language string.

    Example input:
        key="demographics", value={"languages": ["Swedish", "English"], "locations": ["Sweden"]}

    Example output:
        "Languages: Swedish, English. Locations: Sweden."
    """
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float, bool)):
        return str(value)
    if not isinstance(value, dict):
        return json.dumps(value)

    parts = []
    for sub_key, sub_value in value.items():
        label = sub_key.replace("_", " ").title()
        if isinstance(sub_value, list):
            items = [str(v) for v in sub_value if v]
            if items:
                parts.append(f"{label}: {', '.join(items)}")
        elif isinstance(sub_value, str) and sub_value.strip():
            parts.append(f"{label}: {sub_value.strip()}")
        elif isinstance(sub_value, dict):
            nested = flatten_json_section(sub_key, sub_value)
            if nested:
                parts.append(f"{label}: {nested}")

    return ". ".join(parts) + "." if parts else ""


def flatten_json_profile(json_obj):
    """Flatten a JSON profile into a list of (synthetic_key, text) tuples.

    The 'summary' field becomes one thought, each recognized section becomes another.
    """
    results = []

    # Summary first
    summary = json_obj.get("summary", "")
    if isinstance(summary, str) and summary.strip():
        results.append(("profile.summary", summary.strip()))

    # Flattened sections
    for section_key, label in PROFILE_SECTIONS.items():
        section_data = json_obj.get(section_key)
        if section_data is None:
            continue
        text = flatten_json_section(section_key, section_data)
        if text:
            results.append((f"profile.{section_key}", text))

    return results


# ─── Filtering ───────────────────────────────────────────────────────────────


def should_skip_conversation(conv, sync_log, args):
    """Return a skip reason string, or None if the conversation should be processed."""
    dedupe_key = make_dedupe_key(conv["uuid"])

    if dedupe_key in sync_log["ingested_ids"]:
        return "already_imported"

    # Date filtering on CREATED
    created = conv.get("created", "")
    if created:
        try:
            # Handle both "2023-12-24 10:00:00" and "2023-12-24T10:00:00" formats
            for fmt in (
                "%Y-%m-%d %H:%M:%S.%f",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%dT%H:%M:%S.%f",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d",
            ):
                try:
                    conv_date = datetime.strptime(created, fmt).date()
                    break
                except ValueError:
                    continue
            else:
                conv_date = None

            if conv_date:
                if args.after and conv_date < args.after:
                    return "before_date_filter"
                if args.before and conv_date > args.before:
                    return "after_date_filter"
        except Exception:
            pass

    return None


def should_skip_memory(row, sync_log):
    """Return a skip reason string, or None if the memory row should be processed."""
    # Skip deleted or forgotten
    if row.get("IS_DELETED"):
        return "deleted"
    if row.get("IS_FORGOTTEN"):
        return "forgotten"

    # Deduplication key: UUID-based for JSON profiles, key-based for normal rows
    if is_json_profile_row(row):
        dedupe_key = make_dedupe_key("json_profile", row.get("MEMORY_VALUE", "")[:200])
    else:
        dedupe_key = make_dedupe_key(
            row.get("MEMORY_KEY", ""), row.get("FIRST_CREATED_AT", "")
        )

    if dedupe_key in sync_log["ingested_ids"]:
        return "already_imported"

    return None


# ─── LLM Summarization ──────────────────────────────────────────────────────


def summarize_openrouter(title, date_str, answer_text):
    """Summarize a Perplexity Q&A into thoughts using OpenRouter."""
    if not OPENROUTER_API_KEY:
        print("   Warning: Skipping summarization (no OPENROUTER_API_KEY)")
        return []

    truncated = answer_text[:6000]

    resp = http_post_with_retry(
        f"{OPENROUTER_BASE}/chat/completions",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
        },
        body={
            "model": "openai/gpt-4o-mini",
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": SUMMARIZATION_PROMPT},
                {
                    "role": "user",
                    "content": f"Search query: {title}\nDate: {date_str}\n\nPerplexity's answer:\n{truncated}",
                },
            ],
            "temperature": 0,
        },
    )

    if not resp or resp.status_code != 200:
        status = resp.status_code if resp else "no response"
        print(f"   Warning: Summarization failed ({status}), skipping conversation.")
        return []

    try:
        data = resp.json()
        result = json.loads(data["choices"][0]["message"]["content"])
        thoughts = result.get("thoughts", [])
        return [t for t in thoughts if isinstance(t, str) and t.strip()]
    except (KeyError, json.JSONDecodeError, IndexError) as e:
        print(f"   Warning: Failed to parse summarization response: {e}")
        return []


def summarize_ollama(title, date_str, answer_text, model_name="qwen3"):
    """Summarize a Perplexity Q&A using a local Ollama model."""
    truncated = answer_text[:6000]

    prompt = (
        f"{SUMMARIZATION_PROMPT}\n\n"
        f"Search query: {title}\nDate: {date_str}\n\n"
        f"Perplexity's answer:\n{truncated}"
    )

    try:
        resp = requests.post(
            f"{OLLAMA_BASE}/api/generate",
            json={
                "model": model_name,
                "prompt": prompt,
                "stream": False,
                "format": "json",
            },
            timeout=120,
        )
    except requests.RequestException as e:
        print(f"   Warning: Ollama request failed: {e}")
        return []

    if resp.status_code != 200:
        print(f"   Warning: Ollama returned {resp.status_code}")
        return []

    try:
        raw = resp.json().get("response", "")
        result = json.loads(raw)
        thoughts = result.get("thoughts", [])
        return [t for t in thoughts if isinstance(t, str) and t.strip()]
    except (json.JSONDecodeError, KeyError) as e:
        print(f"   Warning: Failed to parse Ollama response: {e}")
        return []


def summarize(title, date_str, answer_text, args):
    """Dispatch to the appropriate summarization backend."""
    if args.model == "ollama":
        return summarize_ollama(title, date_str, answer_text, args.ollama_model)
    return summarize_openrouter(title, date_str, answer_text)


# ─── Embedding Generation ───────────────────────────────────────────────────


def generate_embedding(text):
    """Generate a 1536-dim embedding via OpenRouter (text-embedding-3-small)."""
    truncated = text[:8000]

    resp = http_post_with_retry(
        f"{OPENROUTER_BASE}/embeddings",
        headers={
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
        },
        body={
            "model": "openai/text-embedding-3-small",
            "input": truncated,
        },
    )

    if not resp or resp.status_code != 200:
        status = resp.status_code if resp else "no response"
        print(f"   Warning: Embedding generation failed ({status})")
        return None

    try:
        data = resp.json()
        return data["data"][0]["embedding"]
    except (KeyError, IndexError) as e:
        print(f"   Warning: Failed to parse embedding response: {e}")
        return None


# ─── Ingestion ───────────────────────────────────────────────────────────────


def ingest_thought_supabase(content, metadata_dict, created_at=None, extra_columns=None):
    """Insert a thought directly into Supabase with a generated embedding.

    extra_columns: optional dict of top-level column values to set on the
    inserted row (e.g., {"type": "reference", "source_type": "perplexity-scl"}).
    AJO-local addition so the dashboard's Source dropdown + Type filter chips
    pick up imported rows without a post-process pass.
    """
    embedding = generate_embedding(content)
    if not embedding:
        return {"ok": False, "error": "Failed to generate embedding"}

    body = {
        "content": content,
        "embedding": embedding,
        "metadata": metadata_dict,
    }
    if created_at:
        body["created_at"] = created_at
    if extra_columns:
        for k, v in extra_columns.items():
            if v is not None and k not in body:
                body[k] = v

    resp = http_post_with_retry(
        f"{SUPABASE_URL}/rest/v1/thoughts",
        headers={
            "Content-Type": "application/json",
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Prefer": "return=minimal",
        },
        body=body,
    )

    if not resp:
        return {"ok": False, "error": "No response from Supabase"}

    if resp.status_code not in (200, 201):
        try:
            error_detail = resp.json()
        except ValueError:
            error_detail = resp.text
        return {"ok": False, "error": f"HTTP {resp.status_code}: {error_detail}"}

    return {"ok": True}


# ─── CLI ─────────────────────────────────────────────────────────────────────


def parse_date(s):
    """Parse a YYYY-MM-DD string to a date object."""
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        print(f"Error: Invalid date format '{s}'. Use YYYY-MM-DD.")
        sys.exit(1)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Import Perplexity conversations and memories into Open Brain",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
Examples:
  python import-perplexity.py export.xlsx --dry-run --limit 5
  python import-perplexity.py export.xlsx --type memory
  python import-perplexity.py export.xlsx --type conversations --after 2024-01-01
  python import-perplexity.py export.xlsx --model ollama --ollama-model qwen3
  python import-perplexity.py export.xlsx --report import-report.md""",
    )
    parser.add_argument(
        "xlsx_path",
        help="Path to Perplexity conversations export. Accepts .xlsx (canonical), "
             ".csv (Conversations table), or .json (newer per-Space export with "
             "multi-turn entries[]). When .xlsx, Memory sheet is read from the "
             "same workbook unless --memory overrides.",
    )
    parser.add_argument(
        "--memory",
        type=str,
        default=None,
        metavar="FILE",
        help="Path to Memory file (.xlsx with Memory sheet, or .csv). "
             "Required when the main file is .json or a Conversations-only .csv "
             "and --type includes memory. Auto-resolves to imports/Memory.csv "
             "or imports/Memory.xlsx if those exist alongside the main file.",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Parse and summarize but don't ingest"
    )
    parser.add_argument(
        "--after", type=parse_date, help="Only conversations after YYYY-MM-DD"
    )
    parser.add_argument(
        "--before", type=parse_date, help="Only conversations before YYYY-MM-DD"
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Max items per type to process (0 = unlimited)",
    )
    parser.add_argument(
        "--type",
        choices=["conversations", "memory", "both"],
        default="both",
        help="What to import (default: both)",
    )
    parser.add_argument(
        "--model",
        choices=["openrouter", "ollama"],
        default="openrouter",
        help="LLM backend (default: openrouter)",
    )
    parser.add_argument(
        "--ollama-model", default="qwen3", help="Ollama model name (default: qwen3)"
    )
    parser.add_argument(
        "--verbose", action="store_true", help="Show full content during processing"
    )
    parser.add_argument(
        "--report",
        type=str,
        metavar="FILE",
        help="Write a markdown report of everything imported",
    )
    parser.add_argument(
        "--space-map",
        type=str,
        default="imports/perplexity-spaces.json",
        metavar="FILE",
        help="JSON file mapping Perplexity Space UUIDs to friendly names. "
             "Each named space yields source_type='perplexity-<name>'; "
             "unmapped/no-space conversations stay 'perplexity'.",
    )
    parser.add_argument(
        "--no-prefilter",
        action="store_true",
        help="Disable the heuristic junk prefilter (lets every conversation "
             "reach the LLM judge). Default: prefilter is on.",
    )
    return parser.parse_args()


# ─── Main: Conversations Pipeline ───────────────────────────────────────────


def process_conversations(conversations, sync_log, args, space_map=None):
    """Process and ingest conversations. Returns stats dict."""
    space_map = space_map or {}
    stats = {
        "total": len(conversations),
        "already_imported": 0,
        "space_dropped": 0,
        "heuristic_skipped": 0,
        "heuristic_examples": [],
        "processed": 0,
        "thoughts_generated": 0,
        "ingested": 0,
        "errors": 0,
        "report_entries": [],
    }

    if not conversations:
        return stats

    print(f"\n{'═' * 60}")
    print("Conversations")
    print(f"{'═' * 60}")

    for conv in conversations:
        if args.limit and stats["processed"] >= args.limit:
            break

        skip_reason = should_skip_conversation(conv, sync_log, args)
        if skip_reason:
            if skip_reason == "already_imported":
                stats["already_imported"] += 1
            continue

        # AJO-local: spaces marked action=drop in the spaces map skip
        # every conversation in that space. Use this for noisy/test
        # spaces ("ob test1", "{ THE PROMPT-INATOR }") so the LLM
        # doesn't waste time summarising them.
        space_uuid_check = conv.get("space_uuid")
        if space_uuid_check and space_map.get(space_uuid_check, {}).get("action") == "drop":
            stats["space_dropped"] += 1
            continue

        # AJO-local heuristic prefilter — catches obvious trivia BEFORE we
        # spend an OpenRouter call. Disable with --no-prefilter.
        if not getattr(args, "no_prefilter", False):
            skip, reason = looks_like_junk(conv)
            if skip:
                stats["heuristic_skipped"] += 1
                if len(stats["heuristic_examples"]) < 6:
                    stats["heuristic_examples"].append(
                        f"{(conv.get('title') or '(untitled)')[:60]} -- {reason}"
                    )
                continue

        stats["processed"] += 1
        uuid = conv["uuid"]
        title = conv["title"] or "(untitled)"
        answer_text = conv["answer_text"]

        # Resolve Perplexity Space (from collection_uuid) -> friendly
        # name + per-space classification. Unmapped UUIDs fall back to
        # plain "perplexity" with NULL classification (the worker will
        # decide work/personal during entity extraction).
        space_uuid = conv.get("space_uuid")
        space_info = space_map.get(space_uuid, {}) if space_uuid else {}
        space_name = space_info.get("name") or None
        space_classification = space_info.get("classification") or None
        source_type = f"perplexity-{space_name}" if space_name else "perplexity"

        # Parse date
        created = conv.get("created", "")
        created_iso = parse_timestamp_iso(created)
        date_str = created_iso[:10] if created_iso else ""

        word_count = len(answer_text.split())
        print(f"\n{stats['processed']}. {title}")
        space_tag = f" | space={space_name or space_uuid[:8]+'...' if space_uuid else 'none'}"
        print(f"   {word_count} words | {date_str} | {uuid[:8]}...{space_tag}")

        if not answer_text.strip():
            print("   -> No answer text, skipping")
            continue

        # Summarize
        thoughts = summarize(title, date_str, answer_text, args)
        stats["thoughts_generated"] += len(thoughts)

        if not thoughts:
            print("   -> No thoughts extracted (empty summary)")
            if not args.dry_run:
                dedupe_key = make_dedupe_key(uuid)
                sync_log["ingested_ids"][dedupe_key] = datetime.now(
                    timezone.utc
                ).isoformat()
                save_sync_log(sync_log)
            continue

        if args.verbose or args.dry_run:
            for i, thought in enumerate(thoughts, 1):
                preview = thought if len(thought) <= 200 else thought[:200] + "..."
                print(f"   Thought {i}: {preview}")

        if args.report:
            stats["report_entries"].append(
                {
                    "title": title,
                    "date": date_str,
                    "words": word_count,
                    "thoughts": thoughts,
                }
            )

        if args.dry_run:
            continue

        # Build metadata
        metadata = {
            "source": "perplexity",
            "perplexity_title": title,
            "perplexity_date": date_str,
            "perplexity_uuid": uuid,
        }
        if space_uuid:
            metadata["perplexity_space_uuid"] = space_uuid
        if space_name:
            metadata["perplexity_space"] = space_name
        mode = conv.get("mode")
        if mode:
            metadata["perplexity_mode"] = mode

        # AJO column fields. type='reference' lands these in the right
        # dashboard filter bucket; source_type lets the Source dropdown
        # show "perplexity" and (if mapped) per-space variants.
        # classification: pre-set when the space map has work/personal,
        # otherwise NULL and the worker classifies during entity
        # extraction. Also mirror to metadata.classification (kept in
        # sync with the column by the existing AJO pipeline).
        columns = {"type": "reference", "source_type": source_type}
        if space_classification:
            columns["classification"] = space_classification
            metadata["classification"] = space_classification

        # Ingest thoughts
        all_ok = True
        for i, thought in enumerate(thoughts):
            content = f"[Perplexity: {title} | {date_str}] {thought}"
            result = ingest_thought_supabase(
                content, metadata, created_at=created_iso, extra_columns=columns
            )

            if result.get("ok"):
                stats["ingested"] += 1
                print(f"   -> Thought {i + 1} ingested")
            else:
                stats["errors"] += 1
                all_ok = False
                print(
                    f"   -> ERROR (thought {i + 1}): {result.get('error', 'unknown')}"
                )

            time.sleep(0.2)

        if all_ok:
            dedupe_key = make_dedupe_key(uuid)
            sync_log["ingested_ids"][dedupe_key] = datetime.now(
                timezone.utc
            ).isoformat()
            save_sync_log(sync_log)

    return stats


# ─── Main: Memory Pipeline ──────────────────────────────────────────────────


def process_memory(memories, sync_log, args):
    """Process and ingest memory rows. Returns stats dict."""
    stats = {
        "total": len(memories),
        "already_imported": 0,
        "deleted": 0,
        "forgotten": 0,
        "processed": 0,
        "thoughts_generated": 0,
        "ingested": 0,
        "errors": 0,
        "report_entries": [],
    }

    if not memories:
        return stats

    print(f"\n{'═' * 60}")
    print("Memory")
    print(f"{'═' * 60}")

    for mem in memories:
        if args.limit and stats["processed"] >= args.limit:
            break

        skip_reason = should_skip_memory(mem, sync_log)
        if skip_reason:
            if skip_reason == "already_imported":
                stats["already_imported"] += 1
            elif skip_reason == "deleted":
                stats["deleted"] += 1
            elif skip_reason == "forgotten":
                stats["forgotten"] += 1
            continue

        stats["processed"] += 1

        # Build items to ingest: list of (dedupe_key, synthetic_key, text, metadata, created_at_iso)
        items = []
        is_profile = is_json_profile_row(mem)

        if is_profile:
            try:
                profile_json = json.loads(mem["MEMORY_VALUE"])
            except (json.JSONDecodeError, TypeError):
                print(f"\n   Warning: Failed to parse JSON profile, skipping")
                stats["processed"] -= 1
                continue

            profile_entries = flatten_json_profile(profile_json)
            print(
                f"\n{stats['processed']}. [JSON Profile] ({len(profile_entries)} sections)"
            )

            first_created = mem.get("FIRST_CREATED_AT", "")
            created_iso = parse_timestamp_iso(first_created)

            for synthetic_key, text in profile_entries:
                dedupe_key = make_dedupe_key("json_profile", text[:200])
                meta = {
                    "source": "perplexity_memory",
                    "memory_key": synthetic_key,
                    "memory_confidence": "high",
                    "memory_first_created": first_created,
                    "memory_profile_section": synthetic_key.split(".", 1)[-1],
                }
                items.append((dedupe_key, synthetic_key, text, meta, created_iso))

        else:
            memory_key = mem.get("MEMORY_KEY", "")
            memory_value = mem.get("MEMORY_VALUE", "")
            confidence = mem.get("CONFIDENCE", "")
            first_created = mem.get("FIRST_CREATED_AT", "")
            source_query = mem.get("LAST_UPDATED_QUERY", "")
            created_iso = parse_timestamp_iso(first_created)

            print(f"\n{stats['processed']}. [{memory_key}]")

            dedupe_key = make_dedupe_key(memory_key, first_created)
            meta = {
                "source": "perplexity_memory",
                "memory_key": memory_key,
                "memory_confidence": confidence,
                "memory_first_created": first_created,
                "memory_source_query": source_query,
            }
            items.append((dedupe_key, memory_key, memory_value, meta, created_iso))

        stats["thoughts_generated"] += len(items)

        if args.verbose or args.dry_run:
            for i, (_, key, text, _, _) in enumerate(items, 1):
                preview = text if len(text) <= 200 else text[:200] + "..."
                print(f"   Thought {i} [{key}]: {preview}")

        if args.report:
            if is_profile:
                stats["report_entries"].append(
                    {
                        "label": "[JSON Profile]",
                        "key": "profile",
                        "thoughts": [text for _, _, text, _, _ in items],
                    }
                )
            else:
                stats["report_entries"].append(
                    {
                        "label": mem.get("MEMORY_KEY", ""),
                        "key": mem.get("MEMORY_KEY", ""),
                        "thoughts": [text for _, _, text, _, _ in items],
                    }
                )

        if args.dry_run:
            continue

        # Ingest each item
        all_ok = True
        for i, (dedupe_key, synthetic_key, text, meta, created_iso) in enumerate(items):
            if is_profile:
                section_label = (
                    synthetic_key.split(".", 1)[-1].replace("_", " ").title()
                )
                content = f"[Perplexity Memory: Profile — {section_label}] {text}"
            else:
                content = f"[Perplexity Memory: {synthetic_key}] {text}"

            # AJO column fields. Memory entries are facts about the user
            # (interests, work context, preferences), not Q&A research,
            # so they land as type='reference' with source_type
            # 'perplexity-memory' to distinguish from conversations.
            columns = {"type": "reference", "source_type": "perplexity-memory"}
            result = ingest_thought_supabase(
                content, meta, created_at=created_iso, extra_columns=columns
            )

            if result.get("ok"):
                stats["ingested"] += 1
                print(f"   -> Thought {i + 1} ingested")
            else:
                stats["errors"] += 1
                all_ok = False
                print(
                    f"   -> ERROR (thought {i + 1}): {result.get('error', 'unknown')}"
                )

            time.sleep(0.2)

        if all_ok:
            for dedupe_key, _, _, _, _ in items:
                sync_log["ingested_ids"][dedupe_key] = datetime.now(
                    timezone.utc
                ).isoformat()
            save_sync_log(sync_log)

    return stats


# ─── Main Entry Point ───────────────────────────────────────────────────────


def main():
    args = parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.is_file():
        print(f"Error: File not found: {xlsx_path}")
        sys.exit(1)

    # Validate env vars for live mode
    if not args.dry_run:
        if not SUPABASE_URL:
            print("Error: SUPABASE_URL environment variable required.")
            print(
                "Set it to your Supabase project URL (e.g., https://xxxxx.supabase.co)"
            )
            sys.exit(1)
        if not SUPABASE_SERVICE_ROLE_KEY:
            print("Error: SUPABASE_SERVICE_ROLE_KEY environment variable required.")
            print(
                "This is your Supabase Secret Key (Settings → API → Secret key, starts with sb_secret_)"
            )
            sys.exit(1)
        if not OPENROUTER_API_KEY:
            print(
                "Error: OPENROUTER_API_KEY required for embeddings and summarization."
            )
            print("Get one at https://openrouter.ai/keys")
            sys.exit(1)

    # Warn about missing API key for summarization in dry-run (won't produce summaries)
    if args.dry_run and args.model == "openrouter" and not OPENROUTER_API_KEY:
        print(
            "Note: OPENROUTER_API_KEY not set. Summarization will be skipped in dry-run."
        )
        print("Set the key for a full dry-run preview, or use --model ollama.\n")

    # Display run configuration
    mode = "DRY RUN" if args.dry_run else "LIVE"
    summarize_mode = f"{args.model}"
    if args.model == "ollama":
        summarize_mode += f" ({args.ollama_model})"
    print(f"\n  Mode:        {mode}")
    print(f"  Summarizer:  {summarize_mode}")
    print(f"  Type:        {args.type}")
    if args.after:
        print(f"  After:       {args.after}")
    if args.before:
        print(f"  Before:      {args.before}")
    if args.limit:
        print(f"  Limit:       {args.limit} per type")
    print()

    sync_log = load_sync_log()

    # Load optional space-uuid -> friendly-name map (AJO-local feature).
    space_map = load_space_map(args.space_map)
    if space_map:
        print(f"Loaded {len(space_map)} mapped Perplexity Space(s) from {args.space_map}")

    # Resolve memory file path. If --memory wasn't given:
    #   - .xlsx main file: assume Memory is in same workbook (legacy behaviour)
    #   - otherwise: look for sibling Memory.csv / Memory.xlsx
    main_ext = xlsx_path.suffix.lower()
    if args.memory:
        memory_path = Path(args.memory)
    elif main_ext == ".xlsx":
        memory_path = xlsx_path
    else:
        # auto-discover alongside the main file
        candidates = [xlsx_path.parent / "Memory.csv", xlsx_path.parent / "Memory.xlsx"]
        memory_path = next((c for c in candidates if c.is_file()), None)

    if args.type in ("memory", "both") and not memory_path:
        print(f"Warning: --type {args.type} requested but no Memory file found "
              f"(use --memory PATH).")

    # Process conversations
    conv_stats = None
    if args.type in ("conversations", "both"):
        print(f"Extracting conversations from {xlsx_path}...")
        conversations = extract_conversations(str(xlsx_path))
        print(f"Found {len(conversations)} conversations.")
        conversations.sort(key=lambda c: c.get("created", ""))
        conv_stats = process_conversations(conversations, sync_log, args, space_map=space_map)

    # Process memory
    mem_stats = None
    if args.type in ("memory", "both") and memory_path:
        print(f"\nExtracting memory from {memory_path}...")
        memories = extract_memory_rows(str(memory_path))
        print(f"Found {len(memories)} memory entries.")
        mem_stats = process_memory(memories, sync_log, args)

    # ─── Summary ─────────────────────────────────────────────────────────────

    print(f"\n{'─' * 60}")
    print("Summary:")

    if conv_stats:
        print(f"\n  Conversations:")
        print(f"    Found:              {conv_stats['total']}")
        if conv_stats["already_imported"]:
            print(f"    Already imported:   {conv_stats['already_imported']} (skipped)")
        if conv_stats.get("space_dropped"):
            print(f"    Space dropped:      {conv_stats['space_dropped']} (action=drop in spaces map)")
        if conv_stats.get("heuristic_skipped"):
            print(f"    Heuristic-skipped:  {conv_stats['heuristic_skipped']} (junk filter)")
            for ex in conv_stats.get("heuristic_examples", []):
                print(f"      - {ex}")
        print(f"    Processed:          {conv_stats['processed']}")
        print(f"    Thoughts:           {conv_stats['thoughts_generated']}")
        if not args.dry_run:
            print(f"    Ingested:           {conv_stats['ingested']}")
            print(f"    Errors:             {conv_stats['errors']}")

    if mem_stats:
        print(f"\n  Memory:")
        print(f"    Found:              {mem_stats['total']}")
        if mem_stats["already_imported"]:
            print(f"    Already imported:   {mem_stats['already_imported']} (skipped)")
        if mem_stats["deleted"]:
            print(f"    Deleted:            {mem_stats['deleted']} (skipped)")
        if mem_stats["forgotten"]:
            print(f"    Forgotten:          {mem_stats['forgotten']} (skipped)")
        print(f"    Processed:          {mem_stats['processed']}")
        print(f"    Thoughts:           {mem_stats['thoughts_generated']}")
        if not args.dry_run:
            print(f"    Ingested:           {mem_stats['ingested']}")
            print(f"    Errors:             {mem_stats['errors']}")

    # Cost estimation
    total_thoughts = 0
    total_processed = 0
    if conv_stats:
        total_thoughts += conv_stats["thoughts_generated"]
        total_processed += conv_stats["processed"]
    if mem_stats:
        total_thoughts += mem_stats["thoughts_generated"]
        total_processed += mem_stats["processed"]

    if total_thoughts > 0:
        # Summarization cost (conversations only): gpt-4o-mini via OpenRouter
        # ~$0.15/1M input, ~$0.60/1M output, ~800 tokens in / 200 tokens out per conv
        conv_count = conv_stats["processed"] if conv_stats else 0
        summarize_cost = (conv_count * 800 * 0.15 / 1_000_000) + (
            conv_count * 200 * 0.60 / 1_000_000
        )

        # Embedding cost: $0.02/1M tokens, ~100 tokens per thought
        embedding_cost = total_thoughts * 100 * 0.02 / 1_000_000

        total_cost = summarize_cost + embedding_cost
        print(f"\n  Est. API cost:          ${total_cost:.4f}")
        if conv_count > 0:
            print(f"    Summarization:        ${summarize_cost:.4f}")
        print(f"    Embeddings:           ${embedding_cost:.4f}")

    print(f"{'─' * 60}")

    # Write report
    if args.report:
        _write_report(args.report, conv_stats, mem_stats, args.dry_run)


def _write_report(filepath, conv_stats, mem_stats, dry_run):
    """Write a markdown report of imported data."""
    with open(filepath, "w") as f:
        mode_str = "DRY RUN" if dry_run else "LIVE"
        f.write(f"# Perplexity Import Report ({mode_str})\n\n")
        f.write(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n\n"
        )

        f.write("## Stats\n\n")
        f.write(f"| Metric | Conversations | Memory |\n")
        f.write(f"|--------|--------------|--------|\n")

        conv_total = conv_stats["total"] if conv_stats else 0
        mem_total = mem_stats["total"] if mem_stats else 0
        conv_already = conv_stats["already_imported"] if conv_stats else 0
        mem_already = mem_stats["already_imported"] if mem_stats else 0
        conv_proc = conv_stats["processed"] if conv_stats else 0
        mem_proc = mem_stats["processed"] if mem_stats else 0
        conv_thoughts = conv_stats["thoughts_generated"] if conv_stats else 0
        mem_thoughts = mem_stats["thoughts_generated"] if mem_stats else 0
        conv_ingested = conv_stats["ingested"] if conv_stats else 0
        mem_ingested = mem_stats["ingested"] if mem_stats else 0
        conv_errors = conv_stats["errors"] if conv_stats else 0
        mem_errors = mem_stats["errors"] if mem_stats else 0

        f.write(f"| Found | {conv_total} | {mem_total} |\n")
        f.write(f"| Already imported | {conv_already} | {mem_already} |\n")
        f.write(f"| Processed | {conv_proc} | {mem_proc} |\n")
        f.write(f"| Thoughts | {conv_thoughts} | {mem_thoughts} |\n")
        if not dry_run:
            f.write(f"| Ingested | {conv_ingested} | {mem_ingested} |\n")
            f.write(f"| Errors | {conv_errors} | {mem_errors} |\n")
        f.write("\n")

        # Conversation details
        if conv_stats and conv_stats.get("report_entries"):
            f.write("## Conversations\n\n")
            for entry in conv_stats["report_entries"]:
                f.write(f"### {entry['title']} ({entry['date']})\n\n")
                f.write(f"_{entry['words']} words_\n\n")
                for i, thought in enumerate(entry["thoughts"], 1):
                    f.write(f"{i}. {thought}\n")
                f.write("\n")

        # Memory details
        if mem_stats and mem_stats.get("report_entries"):
            f.write("## Memory\n\n")
            for entry in mem_stats["report_entries"]:
                f.write(f"### {entry['label']}\n\n")
                for i, thought in enumerate(entry["thoughts"], 1):
                    f.write(f"{i}. {thought}\n")
                f.write("\n")

    print(f"\nReport written to {filepath}")


if __name__ == "__main__":
    main()
